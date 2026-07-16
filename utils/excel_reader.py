from openpyxl import load_workbook


def get_registration_data():

    workbook = load_workbook("testdata/registration.xlsx")
    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):

        data.append(
            (
                row[0],  # Name
                row[1],  # Grade
                row[2],  # School
                row[3],  # Email
                row[4],  # File
            )
        )

    return data
