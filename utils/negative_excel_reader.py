from openpyxl import load_workbook


def get_negative_registration_data():

    workbook = load_workbook("testdata/registration_negative.xlsx")

    sheet = workbook.active

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    workbook.close()

    return data
